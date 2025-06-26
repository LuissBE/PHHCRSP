import os
from openpyxl import Workbook

def is_float(value):
    try:
        float(value)  # Try converting the string to float
        return True
    except ValueError:
        return False

def process_text_files():
    # Get the current directory where the script is located
    directory = os.path.dirname(os.path.realpath(__file__))
    
    # Define the paths for the output files
    results_file = os.path.join(directory, "tab_results.txt")
    info_file = os.path.join(directory, "tab_info.txt")
    excel_file = os.path.join(directory, "tab_results.xlsx")
    
    # Create a new workbook and two sheets for Excel output
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Tab_Results"  # First sheet for first lines
    resuts_title = "Algorithm:	Algorithm:	T:	T:	Objective:	Objective:	Status:	Status:	Instance:	Instance:	ObjValue:	ObjValue:	LB:	LB:	Time:	Time:	Threads:	Threads:	Preprocess:	Preprocess:	BinVar:	BinVar:	ContVar:	ContVar:	TVar:	TVar:	NumConst:	NumConst:	TotWorkload:	TotWorkload:	MinWorkload:	MinWorkload:	MaxWorkload:	MaxWorkload:"
    #floats_results = [3, 5, 7, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31, 33]
    cells_results_title = resuts_title.split('\t')
    sheet1.append(cells_results_title)

    sheet2 = wb.create_sheet(title="Tab_Info")  # Second sheet for remaining lines
    info_title = "Algorithm:	Algorithm:	T:	T:	Objective:	Objective:	Instance:	Instance:	Caregiver	Max_day:	Max_day:	day_0:	day_0:	day_1:	day_1:	day_2:	day_2:	day_3:	day_3:	day_4:	day_4:	Total:	Total:	s_day_0:	s_day_0:	s_day_1:	s_day_1:	s_day_2:	s_day_2:	s_day_3:	s_day_3:	s_day_4:	s_day_4:	Wperc:"	
    #floats_info = [3, 5, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32]
    cells_info_title = info_title.split('\t')
    sheet2.append(cells_info_title)
    
    # Open the output text files in write mode
    with open(results_file, 'w') as results, open(info_file, 'w') as info:
        # Loop through all the text files in the current directory
        for filename in os.listdir(directory):
            if filename.endswith('.txt') and filename != 'tab_results.txt' and filename != 'tab_info.txt':  # Avoid output files
                file_path = os.path.join(directory, filename)
                with open(file_path, 'r') as file:
                    lines = file.readlines()
                    for line in lines:
                        if line.strip():  # Non-empty line (ignores lines that are just whitespace)
                            if "Status:" in line:
                                cells_first_non_empty_line = line.split('\t')
                                results.write(line + '\n')
                                 # Write the first non-empty line and the file name to Sheet1
                                for line in cells_first_non_empty_line:
                                    if is_float(line):
                                        line = float(line)
                                sheet1.append(cells_first_non_empty_line)
                            elif "Max_day:" in line:
                                info.write(line + '\n')
                                cells_remaining_content = line.split('\t')
                                for line in cells_remaining_content:
                                    if is_float(line):
                                        line = float(line)
                                row_num = sheet2.max_row + 1
                                Wperc_str = "=20*W" + str(row_num) + "/K" + str(row_num)
                                sheet2.append(cells_remaining_content[:-1] + [Wperc_str])

    # Save the Excel file
    wb.save(excel_file)
    
    print(f"Results saved in {results_file}, {info_file}, and {excel_file}.")

# Run the function
process_text_files()
